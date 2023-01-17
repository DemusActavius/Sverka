from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl import load_workbook
import sys
from openpyxl.styles import PatternFill
import MainWindow

wb1 = ''
wb2 = ''
item1 = ''
item2 = ''

class Sverka(QtWidgets.QMainWindow, QtWidgets.QFileDialog, QtWidgets.QMessageBox,  MainWindow.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pB_Close.pressed.connect(self.app_exit)
        self.pB_open1.pressed.connect(self.open_xls_1)
        self.pB_open2.pressed.connect(self.open_xls_2)
        self.listWidget.clicked.connect(self.get_item1)
        self.listWidget_2.clicked.connect(self.get_item2)
        self.pB_Check.pressed.connect(self.load_xls)

    def __str__(self):
        return self.objectName()

    def app_exit(self):
        app.exit()

    def open_xls_1(self):
        self.label_6.setText("Загружаю список листов")
        global wb1
        wb1 = QFileDialog.getOpenFileName(self, 'Open file', '', '*.*')[0]
        if len(wb1) > 0:
            self.tE_open1.setText(wb1)
            items = self.get_sheet_names(wb1)
            for item in items:
                self.listWidget.addItem(item)
            self.label_6.setText("Загружено")

    def open_xls_2(self):
        self.label_6.setText("Загружаю список листов")
        global wb2
        wb2 = QFileDialog.getOpenFileName(self, 'Open file', '', '*.*')[0]
        if len(wb2) > 0:
            self.tE_open2.setText(wb2)
            items = self.get_sheet_names(wb2)
            for item in items:
                self.listWidget_2.addItem(item)
            self.label_6.setText("Загружено")

    def get_item1(self):
        global item1
        item1 = self.listWidget.currentItem().text()
        return item1

    def get_item2(self):
        global item2
        item2 = self.listWidget_2.currentItem().text()
        return item2

    def get_sheet_names(self, fname):
        wb = load_workbook(fname, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names

    def showDialog(self, showing_text):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Warning)
        msgBox.setText(showing_text)
        msgBox.setWindowTitle("Внимание!!!")
        msgBox.setStandardButtons(QMessageBox.Ok)
        returnValue = msgBox.exec()

    def load_xls(self):
        if len(wb1) > 0:
            if len(wb2) > 0:
                if len(item1) > 0:
                    if len(item2) > 0:
                        self.label_6.setText("Произвожу сверку графиков! Ожидайте.")
                        dif = False
                        i = 0
                        fname_1 = load_workbook(wb1, data_only=True)
                        sheet_1 = fname_1[item1]
                        fname_2 = load_workbook(wb2, data_only=True)
                        sheet_2 = fname_2[item2]
                        for cellObj_1 in sheet_1['B14':'BV570']:
                            for cell_1 in cellObj_1:
                                i += 1
                                self.progressBar.setValue(int((i / 40100) * 100))
                                if cell_1.value == None:
                                    continue
                                if sheet_2.cell(row=cell_1.row, column=cell_1.column).value != cell_1.value:
                                    active_cell = fname_2.active.cell(row=cell_1.row, column=cell_1.column)
                                    active_cell.fill = PatternFill("solid", fgColor="ff0000")
                                    dif = True
                        fname_2.save(wb2)
                        if dif == True:
                            self.label_6.setText("Сверка выполнена! Найдены отличия!")
                        else:
                            self.label_6.setText("Сверка выполнена! Отличия не найдены!")
                    else:
                        self.showDialog("Не выбран лист в графике (послед. изм.)")
                else:
                    self.showDialog("Не выбран лист в графике (утвержден.)")
            else:
                self.showDialog("Не выбран файл с графиком (послед. изм.)")
        else:
            self.showDialog("Не выбран файл с графиком (утвержден.)")


if __name__ == '__main__':
    print(load_workbook.__doc__)
    app = QtWidgets.QApplication(sys.argv)
    window = Sverka()
    window.show()
    app.exec_()

#For conver ui to py in terminal
#pyuic5 XXXX.ui -o XXXXX.py
#
#pyinstaller C:\Users\admin\PycharmProjects\TextToVoice\main.py - w --onefile
